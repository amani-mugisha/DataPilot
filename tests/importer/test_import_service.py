from __future__ import annotations

from io import StringIO, BytesIO

import pandas as pd
from django.test import SimpleTestCase

from apps.importer.services import ImportService


from openpyxl import Workbook


class ImportServiceTests(SimpleTestCase):

    def setUp(self):
        self.service = ImportService()

    def test_supported_formats(self):
        self.assertEqual(
            self.service.supported_formats(),
            [
                "csv",
                "excel_binary",
                "excel_macro",
                "excel_standard",
                "excel_template",
                "excel_template_macro",
            ],
        )

    def test_detect_csv(self):
        detected = self.service.detect(
            "customers.csv"
        )

        self.assertEqual(
            detected.format,
            "csv",
        )

        self.assertEqual(
            detected.extension,
            ".csv",
        )

    def test_validate_csv(self):
        csv_data = StringIO(
            "name,age\n"
            "Amani,20\n"
            "John,25\n"
        )

        detected = self.service.validate(
            file_path=csv_data,
            filename="customers.csv",
        )

        self.assertEqual(
            detected.format,
            "csv",
        )

    def test_read_csv(self):
        csv_data = StringIO(
            "name,age\n"
            "Amani,20\n"
            "John,25\n"
        )

        dataframe, detected = self.service.read(
            file_path=csv_data,
            filename="customers.csv",
        )

        self.assertIsInstance(
            dataframe,
            pd.DataFrame,
        )

        self.assertEqual(
            detected.format,
            "csv",
        )

        self.assertEqual(
            list(dataframe.columns),
            ["name", "age"],
        )

        self.assertEqual(
            len(dataframe),
            2,
        )

    def test_read_preserves_missing_values(self):
        csv_data = StringIO(
            "name,age\n"
            "Amani,20\n"
            "John,\n"
        )

        dataframe, _ = self.service.read(
            file_path=csv_data,
            filename="customers.csv",
        )

        self.assertTrue(
            pd.isna(
                dataframe.loc[1, "age"]
            )
        )

    def test_excel_is_detected_as_supported_format(self):
        detected = self.service.detect(
            "customers.xlsx"
        )

        self.assertEqual(
            detected.format,
            "excel_standard",
        )

        self.assertEqual(
            detected.extension,
            ".xlsx",
        )

    def test_reads_xlsx(self):
        from io import BytesIO

        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active

        worksheet.append(["name", "age"])
        worksheet.append(["Amani", 20])
        worksheet.append(["John", 25])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        dataframe, detected = self.service.read(
            file_path=buffer,
            filename="customers.xlsx",
        )

        self.assertEqual(
            detected.format,
            "excel_standard",
        )

        self.assertEqual(
            list(dataframe.columns),
            ["name", "age"],
        )

        self.assertEqual(
            len(dataframe),
            2,
        )

    def test_rejects_empty_csv(self):
        csv_data = StringIO("")

        with self.assertRaisesMessage(
            ValueError,
            "The CSV file is empty.",
        ):
            self.service.read(
                file_path=csv_data,
                filename="customers.csv",
            )

    def test_every_reader_is_a_supported_format(self):
        """
        Every registered reader must correspond to a supported
        DataPilot format.
        """
        from apps.importer.formats import SUPPORTED_FORMATS

        supported_formats = set(
            SUPPORTED_FORMATS.values()
        )

        self.assertTrue(
            set(self.service.readers).issubset(
                supported_formats
            )
        )

    def test_every_validator_is_a_supported_format(self):
        """
        Every registered validator must correspond to a supported
        DataPilot format.
        """
        from apps.importer.formats import SUPPORTED_FORMATS

        supported_formats = set(
            SUPPORTED_FORMATS.values()
        )

        self.assertTrue(
            set(self.service.validators).issubset(
                supported_formats
            )
        )

    def test_registered_csv_reader(self):
        self.assertIn(
            "csv",
            self.service.readers,
        )

        self.assertEqual(
            self.service.readers["csv"].format_name,
            "csv",
        )

    def test_registered_csv_validator(self):
        self.assertIn(
            "csv",
            self.service.validators,
        )

        self.assertEqual(
            self.service.validators["csv"].format_name,
            "csv",
        )

    def test_validation_only_formats_do_not_require_reader(self):
        """
        Some formats may be detectable and validatable without
        being importable into a DataFrame.
        """
        self.assertIn(
            "excel_addin",
            self.service.validators,
        )

        self.assertNotIn(
            "excel_addin",
            self.service.readers,
        )

    def _create_excel_workbook(self) -> BytesIO:
        workbook = Workbook()

        customers = workbook.active
        customers.title = "Customers"
        customers.append(["name", "age"])
        customers.append(["Amani", 20])
        customers.append(["John", 25])

        orders = workbook.create_sheet("Orders")
        orders.append(["product", "price"])
        orders.append(["Laptop", 1000])
        orders.append(["Phone", 500])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return buffer

    def test_list_excel_sheets(self):
        workbook = self._create_excel_workbook()

        sheets = self.service.list_sheets(
            file_path=workbook,
            filename="customers.xlsx",
        )

        self.assertEqual(
            sheets,
            ["Customers", "Orders"],
        )

    def test_read_excel_sheet_by_name(self):
        workbook = self._create_excel_workbook()

        dataframe, detected = self.service.read(
            file_path=workbook,
            filename="customers.xlsx",
            sheet_name="Orders",
        )

        self.assertEqual(
            detected.format,
            "excel_standard",
        )

        self.assertEqual(
            list(dataframe.columns),
            ["product", "price"],
        )

        self.assertEqual(
            len(dataframe),
            2,
        )

    def test_read_excel_sheet_by_index(self):
        workbook = self._create_excel_workbook()

        dataframe, detected = self.service.read(
            file_path=workbook,
            filename="customers.xlsx",
            sheet_name=1,
        )

        self.assertEqual(
            detected.format,
            "excel_standard",
        )

        self.assertEqual(
            list(dataframe.columns),
            ["product", "price"],
        )

    def test_list_sheets_rejects_csv(self):
        csv_data = StringIO(
            "name,age\n"
            "Amani,20\n"
        )

        with self.assertRaisesMessage(
            ValueError,
            "Sheet listing is only supported for Excel files.",
        ):
            self.service.list_sheets(
                file_path=csv_data,
                filename="customers.csv",
            )

    def test_read_excel_invalid_sheet(self):
        workbook = self._create_excel_workbook()

        with self.assertRaisesMessage(
            ValueError,
            "Could not read Excel file:",
        ):
            self.service.read(
                file_path=workbook,
                filename="customers.xlsx",
                sheet_name="MissingSheet",
            )

    def test_reads_xlsx_through_complete_import_pipeline(self):
        workbook = self._create_excel_workbook()

        result = self.service.read(
            file_path=workbook,
            filename="customers.xlsx",
        )

        self.assertEqual(
            result.filename,
            "customers.xlsx",
        )

        self.assertEqual(
            result.extension,
            ".xlsx",
        )

        self.assertEqual(
            result.format,
            "excel_standard",
        )

        self.assertIsInstance(
            result.dataframe,
            pd.DataFrame,
        )

        self.assertEqual(
            list(result.dataframe.columns),
            ["name", "age"],
        )